"""
Account Scorer — Salesforce Account Prioritization Tool
Built for Salesforce sales teams selling the full suite to Canadian companies.
"""

import streamlit as st
import pandas as pd
import json
import os
import re
import time
from datetime import datetime

st.set_page_config(
    page_title="Account Scorer — Salesforce",
    page_icon="☁️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS for polished look ────────────────────────────────────
st.markdown("""
<style>
/* ── Global typography ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* ── Hide Streamlit branding ── */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}

/* ── Sidebar polish ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #032D60 0%, #0176D3 100%);
}
[data-testid="stSidebar"] * {
    color: #FFFFFF !important;
}
[data-testid="stSidebar"] .stTextInput label,
[data-testid="stSidebar"] .stFileUploader label,
[data-testid="stSidebar"] .stSlider label {
    color: rgba(255,255,255,0.85) !important;
    font-weight: 500;
}
[data-testid="stSidebar"] .stTextInput input {
    background: rgba(255,255,255,0.15) !important;
    border: 1px solid rgba(255,255,255,0.3) !important;
    color: white !important;
}
[data-testid="stSidebar"] .stButton > button {
    background: #04A5D9 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    padding: 0.6rem 1.2rem !important;
    width: 100%;
    transition: all 0.2s ease;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #0288D1 !important;
    box-shadow: 0 4px 12px rgba(1, 118, 211, 0.4);
}
[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.2) !important;
}
[data-testid="stSidebar"] .stMarkdown p,
[data-testid="stSidebar"] .stMarkdown li {
    color: rgba(255,255,255,0.9) !important;
}

/* ── Main content polish ── */
.main .block-container {
    padding-top: 2rem;
    max-width: 1400px;
}

/* ── Metric cards ── */
[data-testid="stMetric"] {
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 1rem 1.25rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
}
[data-testid="stMetric"] label {
    color: #64748B !important;
    font-size: 0.8rem !important;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: #0F172A !important;
    font-weight: 700 !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 0;
    background: #F1F5F9;
    border-radius: 10px;
    padding: 4px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    padding: 8px 20px;
    font-weight: 500;
    color: #475569;
}
.stTabs [aria-selected="true"] {
    background: white !important;
    color: #0176D3 !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
}

/* ── Tables ── */
.stDataFrame {
    border-radius: 12px;
    overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
}

/* ── Expanders ── */
.streamlit-expanderHeader {
    background: #F8FAFC;
    border-radius: 10px !important;
    font-weight: 600;
}

/* ── Progress bar ── */
.stProgress > div > div > div > div {
    background: linear-gradient(90deg, #0176D3, #04A5D9) !important;
    border-radius: 10px;
}

/* ── Download buttons ── */
.stDownloadButton > button {
    background: white !important;
    color: #0176D3 !important;
    border: 2px solid #0176D3 !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    transition: all 0.2s ease;
}
.stDownloadButton > button:hover {
    background: #0176D3 !important;
    color: white !important;
}

/* ── Score badge styling ── */
.score-high { color: #15803D; font-weight: 700; }
.score-med  { color: #B45309; font-weight: 700; }
.score-low  { color: #DC2626; font-weight: 700; }

/* ── Info/Success/Warning boxes ── */
.stAlert {
    border-radius: 10px !important;
}
</style>
""", unsafe_allow_html=True)

# ── Helpers ─────────────────────────────────────────────────────────

# Words that indicate scraped navigation text (not real content)
_NAV_JUNK = {
    "news", "technology", "launch", "government", "military", "finance",
    "connectivity", "interactive", "features", "about", "contact", "menu",
    "search", "login", "sign in", "subscribe", "newsletter", "home",
    "privacy", "cookie", "terms", "careers", "blog", "press", "investors",
    "support", "help", "faq", "resources", "products", "solutions",
    "services", "partners", "company", "overview", "skip to content",
}


def _clean_summary(text):
    """Strip navigation/menu junk from Exa summaries."""
    if not text:
        return ""
    # Split on sentence boundaries (period followed by capital letter)
    # to handle single-line mixed content without breaking URLs/abbreviations
    text = re.sub(r"\.(\s+)([A-Z])", r".\n\2", text)
    lines = text.split("\n")
    clean = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Skip lines that are mostly nav keywords
        words = [w.strip().lower() for w in re.split(r"[|,/]", line)]
        if len(words) >= 3 and sum(1 for w in words if w in _NAV_JUNK) >= 2:
            continue
        # Skip very short lines that match nav patterns
        if len(line) < 30 and line.lower().strip() in _NAV_JUNK:
            continue
        # Skip lines that are just short fragments with underscores (truncated nav)
        if line.endswith("_") and len(line) < 50:
            continue
        clean.append(line)
    return " ".join(clean).strip()


def _build_markdown_table(rows, columns):
    """Build a compact markdown table from a list of dicts."""
    if not rows:
        return ""
    header = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join("---" for _ in columns) + " |"
    body_lines = []
    for row in rows:
        cells = []
        for col in columns:
            val = str(row.get(col, "")).replace("\n", " ").replace("|", "\\|")
            cells.append(val)
        body_lines.append("| " + " | ".join(cells) + " |")
    return "\n".join([header, separator] + body_lines)


# ── Salesforce Solution Mapping ─────────────────────────────────────
# Maps detected tech categories/products to the Salesforce solution that replaces them

# Categories that are relevant to Salesforce selling
SALESFORCE_RELEVANT_CATEGORIES = {
    "CRM", "Marketing", "Sales Tools", "Service/Support", "E-commerce",
    "Analytics", "CDP", "Personalization", "Email", "Advertising",
    "Retargeting", "Segmentation", "Cloud", "Hosting", "DMS",
    "Containers", "Monitoring", "DevOps",
}

# Map specific products to Salesforce solutions
SALESFORCE_SOLUTION_MAP = {
    # CRM
    "HubSpot": "Sales Cloud",
    "HubSpot (via SPF)": "Sales Cloud",
    "HubSpot Chat": "Service Cloud (Live Agent)",
    "HubSpot CMS Hub": "Experience Cloud",
    "Zoho CRM": "Sales Cloud",
    "Microsoft Dynamics": "Sales Cloud",
    "Pipedrive": "Sales Cloud",
    "Freshsales": "Sales Cloud",
    "Monday Sales CRM": "Sales Cloud",
    "Qualified": "Sales Cloud (Einstein)",
    "Salesforce": "✅ Already using",
    "Salesforce (via SPF)": "✅ Already using",
    "Salesforce Pardot": "✅ Already using (Marketing Cloud Account Engagement)",
    "Salesforce Pardot (via SPF)": "✅ Already using (MCAE)",
    # Marketing
    "Marketo": "Marketing Cloud Account Engagement",
    "Mailchimp": "Marketing Cloud",
    "ActiveCampaign": "Marketing Cloud",
    "Constant Contact": "Marketing Cloud",
    "Braze": "Marketing Cloud",
    "Klaviyo": "Marketing Cloud",
    "6sense": "Marketing Cloud (Account Engagement)",
    "PathFactory": "Marketing Cloud",
    "HubSpot Analytics": "Marketing Cloud Intelligence",
    "Demandbase": "Marketing Cloud (ABM)",
    # Service
    "Zendesk": "Service Cloud",
    "Zendesk (via support)": "Service Cloud",
    "Freshdesk": "Service Cloud",
    "Freshdesk (via support)": "Service Cloud",
    "Intercom": "Service Cloud",
    "Intercom (via SPF)": "Service Cloud",
    "Drift": "Service Cloud (Einstein Bots)",
    "LiveChat": "Service Cloud (Live Agent)",
    "Tidio": "Service Cloud (Live Agent)",
    "Crisp": "Service Cloud (Live Agent)",
    # Commerce
    "Shopify": "Commerce Cloud",
    "Shopify (via store)": "Commerce Cloud",
    "Squarespace Commerce": "Commerce Cloud",
    "WooCommerce": "Commerce Cloud",
    "Magento": "Commerce Cloud",
    "BigCommerce": "Commerce Cloud",
    "commercetools": "Commerce Cloud",
    "VTEX": "Commerce Cloud",
    "Elastic Path": "Commerce Cloud",
    # Analytics
    "Google Analytics": "CRM Analytics (Tableau)",
    "Adobe Analytics": "CRM Analytics (Tableau)",
    "Hotjar": "CRM Analytics",
    "Mixpanel": "CRM Analytics",
    "Amplitude": "CRM Analytics",
    "VWO": "CRM Analytics",
    "Smartlook": "CRM Analytics",
    "Mouse Flow": "CRM Analytics",
    "Facebook Pixel": "Marketing Cloud (Data Cloud)",
    "Linkedin Insight Tag": "Marketing Cloud (Data Cloud)",
    # CDP / Personalization
    "Segment": "Data Cloud",
    "mParticle": "Data Cloud",
    "Tealium": "Data Cloud",
    "LiveIntent": "Data Cloud",
    "Datadog": "—",
    # Sales tools
    "SalesLoft": "Sales Cloud (Sales Engagement)",
    "Outreach": "Sales Cloud (Sales Engagement)",
    "Gong": "Sales Cloud (Einstein Conversation Insights)",
    # Cloud / Hosting (relevant for AWS Marketplace)
    "Amazon ALB": "Available via AWS Marketplace",
    "AWS": "Available via AWS Marketplace",
    "Cloudflare": "—",
    "Cloudflare Bot Management": "—",
    # Email
    "SendGrid": "Marketing Cloud",
    "Google Workspace": "Salesforce Inbox",
    "Microsoft 365": "Salesforce Inbox",
    # ITSM (Agentforce IT Service)
    "ServiceNow": "Agentforce IT Service",
    "BMC Helix": "Agentforce IT Service",
    "BMC Remedy": "Agentforce IT Service",
    "Freshservice": "Agentforce IT Service",
    "ManageEngine": "Agentforce IT Service",
    "SysAid": "Agentforce IT Service",
    "Ivanti": "Agentforce IT Service",
    # HR / HCM (Employee Service)
    "Workday": "Employee Service",
    "BambooHR": "Employee Service",
    "ADP": "Employee Service",
    "UKG": "Employee Service",
    "Ceridian": "Employee Service",
    "Dayforce": "Employee Service",
    "SAP SuccessFactors": "Employee Service",
    "Rippling": "Employee Service",
    "Personio": "Employee Service",
    # Data Warehouse / Lakehouse (Data Cloud / Data 360)
    "Snowflake": "Data Cloud (Data 360)",
    "Databricks": "Data Cloud (Data 360)",
    "BigQuery": "Data Cloud (Data 360)",
    "Redshift": "Data Cloud (Data 360)",
    "Azure Synapse": "Data Cloud (Data 360)",
    "Teradata": "Data Cloud (Data 360)",
    # AI / LLM (Salesforce AI / Agentforce)
    "OpenAI": "Salesforce AI (Agentforce)",
    "ChatGPT": "Salesforce AI (Agentforce)",
    "Anthropic": "Salesforce AI (Agentforce)",
    "Cohere": "Salesforce AI (Agentforce)",
    "Google Gemini": "Salesforce AI (Agentforce)",
    "Amazon Bedrock": "Salesforce AI (Agentforce 360 for AWS)",
    # iPaaS / Integration (MuleSoft)
    "Boomi": "MuleSoft",
    "Informatica": "MuleSoft / Data Cloud",
    "Workato": "MuleSoft",
    "Tray.io": "MuleSoft",
    "Celigo": "MuleSoft",
    "SnapLogic": "MuleSoft",
    "Jitterbit": "MuleSoft",
    "Zapier": "MuleSoft / Flow",
    # Collaboration (Slack)
    "Microsoft Teams": "Slack",
    "Google Chat": "Slack",
    "Zoom": "Slack",
    "Webex": "Slack",
    # CCaaS (Service Cloud Voice)
    "Five9": "Service Cloud Voice",
    "NICE CXone": "Service Cloud Voice",
    "Genesys": "Service Cloud Voice",
    "Talkdesk": "Service Cloud Voice",
    "RingCentral": "Slack + Service Cloud Voice",
    "Dialpad": "Slack + Service Cloud Voice",
    "Aircall": "Service Cloud Voice",
    # Data Backup (Own / Salesforce Backup)
    "Odaseva": "Salesforce Backup (Own)",
    "Veeam": "Salesforce Backup (Own)",
    "Druva": "Salesforce Backup (Own)",
    "Commvault": "Salesforce Backup (Own)",
    "OwnBackup": "✅ Already using (Own)",
    # Data Security (Shield)
    "Varonis": "Shield + Data Mask",
    "BigID": "Shield + Privacy Center",
    "OneTrust": "Shield + Privacy Center",
    # Field Service
    "ServiceMax": "Field Service Lightning",
    "IFS": "Field Service Lightning",
    # ERP (MuleSoft integration)
    "SAP": "Revenue Cloud / MuleSoft integration",
    "Oracle ERP": "Revenue Cloud / MuleSoft integration",
    "NetSuite": "Revenue Cloud / MuleSoft integration",
    "Sage": "Revenue Cloud / MuleSoft integration",
    "Sage Intacct": "Revenue Cloud / MuleSoft integration",
    "Infor": "Revenue Cloud / MuleSoft integration",
    "QuickBooks": "Revenue Cloud / MuleSoft integration",
    # CDP
    "Segment": "Data Cloud",
    "mParticle": "Data Cloud",
    "Tealium": "Data Cloud",
    "Treasure Data": "Data Cloud",
    "BlueConic": "Data Cloud",
    # Sustainability (Net Zero Cloud)
    "Persefoni": "Net Zero Cloud",
    "Watershed": "Net Zero Cloud",
    # Customer Success
    "Gainsight": "Service Cloud + Data Cloud",
    "Totango": "Service Cloud + Data Cloud",
    "ChurnZero": "Service Cloud + Data Cloud",
    # Sales Enablement
    "Highspot": "Sales Cloud (Enablement)",
    "Seismic": "Sales Cloud (Enablement)",
    "Showpad": "Sales Cloud (Enablement)",
    # Revenue Intelligence
    "Clari": "Sales Cloud (Revenue Intelligence)",
    "People.ai": "Sales Cloud (Einstein Activity Capture)",
    # CPQ / Billing
    "Zuora": "Revenue Cloud (CPQ + Billing)",
    "Chargebee": "Revenue Cloud",
    "DealHub": "Revenue Cloud (CPQ)",
    # Community / Experience / Portal
    "Khoros": "Experience Cloud",
    "Higher Logic": "Experience Cloud",
    "Lithium": "Experience Cloud",
    "Liferay": "Experience Cloud",
    "Adobe Experience Manager": "Experience Cloud",
    "Sitecore": "Experience Cloud",
    "Contentful": "Experience Cloud",
    "Contentstack": "Experience Cloud",
    "Acquia": "Experience Cloud",
    # RPA (MuleSoft RPA)
    "UiPath": "MuleSoft RPA",
    "Automation Anywhere": "MuleSoft RPA",
    "Blue Prism": "MuleSoft RPA",
    # Loyalty Management
    "Antavo": "Loyalty Management",
    "Annex Cloud": "Loyalty Management",
    "Yotpo Loyalty": "Loyalty Management",
    "Smile.io": "Loyalty Management",
    "LoyaltyLion": "Loyalty Management",
    # Enterprise Search (Slack)
    "Glean": "Slack (Enterprise Search)",
    "Coveo": "Slack (Enterprise Search)",
    "Algolia": "Slack (Enterprise Search)",
    "Lucidworks": "Slack (Enterprise Search)",
    # Financial Services (Financial Services Cloud)
    "nCino": "Financial Services Cloud",
    "Finastra": "Financial Services Cloud",
    "Temenos": "Financial Services Cloud",
    "Fiserv": "Financial Services Cloud",
    "Jack Henry": "Financial Services Cloud",
    "Verint": "Financial Services Cloud",
    # Healthcare (Health Cloud)
    "Epic": "Health Cloud",
    "Cerner": "Health Cloud",
    "Oracle Health": "Health Cloud",
    "Veeva": "Health Cloud",
    "Medidata": "Health Cloud",
    "athenahealth": "Health Cloud",
    # Nonprofit (Nonprofit Cloud)
    "Blackbaud": "Nonprofit Cloud",
    "Bloomerang": "Nonprofit Cloud",
    "Neon CRM": "Nonprofit Cloud",
    "Raiser's Edge": "Nonprofit Cloud",
    "Classy": "Nonprofit Cloud",
    # Data Governance / MDM (Informatica)
    "Collibra": "Data Cloud (Informatica)",
    "Atlan": "Data Cloud (Informatica)",
    "Alation": "Data Cloud (Informatica)",
    "Reltio": "Data Cloud (Informatica)",
}

# Categories to completely hide (not relevant to SF selling)
HIDE_CATEGORIES = {
    "A/B Testing", "Font scripts", "JavaScript frameworks", "JavaScript libraries",
    "UI frameworks", "Web frameworks", "Web servers", "Programming languages",
    "Miscellaneous", "Privacy", "Accessibility", "Security", "CDN",
    "JavaScript graphics", "Tag Management", "Video players", "Maps",
    "Payments", "Databases", "Rich text editors",
    "SSL/TLS certificate authorities", "Caching", "Cookie compliance",
    "Feature management", "Translation", "Photo galleries",
    "Live chat",  # generic category — specific chat tools like Drift/Intercom have their own
}


def _get_sf_solution(tech_name, category):
    """Get the Salesforce solution that could replace this tech."""
    # Check exact product match first
    if tech_name in SALESFORCE_SOLUTION_MAP:
        return SALESFORCE_SOLUTION_MAP[tech_name]

    # Check partial matches
    name_lower = tech_name.lower()
    for key, solution in SALESFORCE_SOLUTION_MAP.items():
        if key.lower() in name_lower:
            return solution

    # Default by category
    category_defaults = {
        "CRM": "Sales Cloud",
        "Marketing": "Marketing Cloud",
        "Service/Support": "Service Cloud",
        "E-commerce": "Commerce Cloud",
        "Analytics": "CRM Analytics (Tableau)",
        "CDP": "Data Cloud",
        "Sales Tools": "Sales Cloud",
        "Personalization": "Data Cloud",
        "Email": "Marketing Cloud",
        "Advertising": "Marketing Cloud",
        "Retargeting": "Marketing Cloud",
        "Segmentation": "Data Cloud",
        "Cloud": "Available via AWS Marketplace",
        "DMS": "—",
        "Monitoring": "—",
    }
    return category_defaults.get(category, "—")


# ── Session State Init ──────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = None
if "running" not in st.session_state:
    st.session_state.running = False
if "progress" not in st.session_state:
    st.session_state.progress = 0
if "log" not in st.session_state:
    st.session_state.log = []


# ── Sidebar ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="text-align: center; padding: 0.5rem 0 1rem 0;">
        <img src="https://upload.wikimedia.org/wikipedia/commons/thumb/f/f9/Salesforce.com_logo.svg/200px-Salesforce.com_logo.svg.png"
             width="140" style="filter: brightness(0) invert(1); opacity: 0.95;" />
        <h2 style="margin: 0.5rem 0 0 0; font-size: 1.4rem; font-weight: 700; letter-spacing: -0.02em;">
            Account Scorer
        </h2>
        <p style="margin: 0; font-size: 0.85rem; opacity: 0.7;">
            Prioritize accounts. Find opportunities.
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    # Load saved key: Streamlit secrets > env var > local file
    default_key = ""
    try:
        default_key = st.secrets.get("EXA_API_KEY", "")
    except Exception:
        pass
    if not default_key:
        default_key = os.environ.get("EXA_API_KEY", "")
    key_file = os.path.join(os.path.dirname(__file__), ".exa_key")
    if not default_key and os.path.exists(key_file):
        with open(key_file) as f:
            default_key = f.read().strip()

    exa_key = st.text_input(
        "Exa API Key",
        value=default_key,
        type="password",
        help="Get a free key at exa.ai — $10 free credits, no CC required",
    )

    # Save key for next session
    if exa_key and exa_key != default_key:
        with open(key_file, "w") as f:
            f.write(exa_key)

    if not exa_key:
        st.warning("Enter your Exa API key to enable enrichment")

    st.divider()
    st.markdown("##### Enrichment Layers")
    do_website = st.checkbox("Website Discovery", value=True)
    do_tech_stack = st.checkbox("Tech Stack Analysis", value=True)
    do_job_postings = st.checkbox("Job Posting Tech Scan", value=True)
    do_company_intel = st.checkbox("Company Intel", value=True)
    do_news = st.checkbox("News & Events", value=True)
    do_executives = st.checkbox("Executive Changes", value=True)

    st.divider()
    st.markdown("##### Scoring Weights")
    w_tech = st.slider("Tech Stack Fit", 0, 100, 30)
    w_events = st.slider("Compelling Events", 0, 100, 25)
    w_execs = st.slider("New Executives", 0, 100, 20)
    w_size = st.slider("Company Size", 0, 100, 15)
    w_access = st.slider("Accessibility", 0, 100, 10)
    st.caption("AOV multiplier: $600k+ = 1.25x · $200k-600k = 1.15x · $100k-200k = 1.1x · $50k-100k = 1.05x · $10k-50k = 1.0x · <$10k = 0.9x · $0 = 0.75x")


# ── Main Area ───────────────────────────────────────────────────────
st.markdown("""
<div style="margin-bottom: 1.5rem;">
    <h1 style="margin: 0; font-size: 2rem; font-weight: 700; color: #0F172A; letter-spacing: -0.03em;">
        ☁️ Account Scorer
    </h1>
    <p style="margin: 0.25rem 0 0 0; font-size: 1.05rem; color: #64748B;">
        Upload your account list to enrich, score, and rank accounts for Salesforce opportunities.
    </p>
</div>
""", unsafe_allow_html=True)

# ── File Upload ─────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload Account List",
    type=["xlsx", "csv"],
    help="Excel or CSV with an 'Account Name' column.",
)

# ── Demo Mode — load saved results ─────────────────────────────────
demo_file = os.path.join(os.path.dirname(__file__), "demo_results.json")
if not uploaded and os.path.exists(demo_file) and not st.session_state.results:
    if st.button("🎯 Load Demo Results", help="View previously enriched accounts"):
        with open(demo_file) as f:
            demo_data = json.load(f)
        # Score each result if not already scored
        from enrichment import score_company
        default_weights = {
            "tech_weight": 30,
            "events_weight": 25,
            "execs_weight": 15,
            "size_weight": 15,
            "accessibility_weight": 15,
        }
        scored = []
        for r in demo_data:
            if not r.get("score"):
                r = score_company(r, default_weights)
            scored.append(r)
        st.session_state.results = scored
        st.rerun()

if uploaded:
    # Parse the file
    try:
        def _has_real_headers(df):
            """Check if the first row looks like real column headers vs data."""
            cols = [str(c) for c in df.columns]
            # Known header patterns — must match as whole words in short labels
            header_patterns = {
                "account name", "account", "company name", "company",
                "status", "aov band", "aov", "headcount", "head count",
                "website", "notes", "tier", "industry", "owner",
                "region", "revenue", "employee count", "employees",
            }
            real_header_hits = 0
            for c in cols:
                cl = str(c).lower().strip()
                # Skip "Unnamed:" columns — they indicate missing headers
                if cl.startswith("unnamed:"):
                    continue
                # Exact or close match against known header labels
                if cl in header_patterns or any(cl.startswith(hp) for hp in header_patterns):
                    real_header_hits += 1
            if real_header_hits >= 2:
                return True

            # Count data-like indicators
            data_indicators = 0
            for c in cols:
                cs = str(c)
                if cs.startswith("$") or cs.startswith("http") or cs.startswith("www."):
                    data_indicators += 1
                if "Unnamed:" in cs:
                    data_indicators += 1
                # Integer or float column names
                if isinstance(c, (int, float)):
                    data_indicators += 1
            if data_indicators > len(cols) * 0.25:
                return False
            return True

        if uploaded.name.endswith(".csv"):
            df = pd.read_csv(uploaded)
            if not _has_real_headers(df):
                uploaded.seek(0)
                df = pd.read_csv(uploaded, header=None)
            df["_team_member"] = "Uploaded"
            all_companies = df
        else:
            xls = pd.ExcelFile(uploaded)
            frames = []
            for sheet in xls.sheet_names:
                sdf = pd.read_excel(xls, sheet_name=sheet)
                if not _has_real_headers(sdf):
                    sdf = pd.read_excel(xls, sheet_name=sheet, header=None)
                sdf["_team_member"] = sheet
                frames.append(sdf)
            all_companies = pd.concat(frames, ignore_index=True)

        # ── Detect headerless files ──────────────────────────────────
        # If all column names are integers, the file has no header row.
        # Try to auto-detect columns by inspecting cell content.
        original_cols = [c for c in all_companies.columns if c != "_team_member"]
        if all(isinstance(c, (int, float)) for c in original_cols):
            # Sniff each column to guess what it contains
            _col_map = {}
            for col in all_companies.columns:
                sample = all_companies[col].dropna().head(5)
                if sample.empty:
                    continue
                sample_str = sample.astype(str)
                joined = " ".join(sample_str).lower()

                # Account name: first string column that isn't a dollar value/URL/status
                if "name" not in _col_map and sample.dtype == object:
                    first_val = str(sample.iloc[0]).strip().lower()
                    if (not first_val.startswith("$") and
                        not first_val.startswith("http") and
                        first_val not in ("can prospect", "cannot prospect", "nan") and
                        "linkedin.com" not in first_val and
                        "tier " not in first_val and
                        len(first_val) > 2):
                        _col_map["name"] = col

                # AOV band: contains dollar values like "$600k-1M"
                if "aov" not in _col_map and sample.dtype == object:
                    if sample_str.str.contains(r'^\$', regex=True).mean() > 0.5:
                        _col_map["aov"] = col

                # Status: contains "Can Prospect"
                if "status" not in _col_map and sample.dtype == object:
                    if sample_str.str.lower().str.contains("prospect").any():
                        _col_map["status"] = col

                # Website: contains URLs (non-LinkedIn)
                if "website" not in _col_map and sample.dtype == object:
                    if (sample_str.str.contains(r'^https?://|^www\.', regex=True).mean() > 0.5 and
                        not sample_str.str.contains("linkedin", case=False).any() and
                        not sample_str.str.contains("force.com", case=False).any()):
                        _col_map["website"] = col

                # Headcount: contains employee count text like "201-500 employees"
                if "headcount" not in _col_map and sample.dtype == object:
                    if sample_str.str.contains("employee", case=False).any():
                        _col_map["headcount"] = col

            # Rename detected columns for consistent downstream use
            rename_map = {}
            if "name" in _col_map:
                rename_map[_col_map["name"]] = "Account Name"
            if "aov" in _col_map:
                rename_map[_col_map["aov"]] = "AOV Band"
            if "status" in _col_map:
                rename_map[_col_map["status"]] = "Status"
            if "website" in _col_map:
                rename_map[_col_map["website"]] = "Website"
            if "headcount" in _col_map:
                rename_map[_col_map["headcount"]] = "Headcount"
            all_companies = all_companies.rename(columns=rename_map)

            # Drop columns that are still integers (unmapped) to keep it clean
            # but keep them in case they're useful
            st.caption(f"Auto-detected columns: {', '.join(rename_map.values())}")

        # ── Find columns by header name ───────────────────────────────
        def _col_lower(col):
            """Safely lowercase a column name (handles int column names)."""
            return str(col).lower()

        name_col = None
        for col in all_companies.columns:
            cl = _col_lower(col)
            if "account" in cl and "name" in cl:
                name_col = col
                break
        if not name_col:
            for col in all_companies.columns:
                cl = _col_lower(col)
                if "company" in cl or cl == "name":
                    name_col = col
                    break
        if not name_col:
            # Last resort: first string column
            for col in all_companies.columns:
                if all_companies[col].dtype == object:
                    name_col = col
                    break
        if not name_col:
            name_col = all_companies.columns[0]

        # Find status column
        status_col = None
        for col in all_companies.columns:
            if "status" in _col_lower(col) or _col_lower(col) == "status":
                status_col = col
                break

        # Filter to Can Prospect only
        if status_col:
            before = len(all_companies)
            prospectable = all_companies[
                all_companies[status_col].astype(str).str.lower().str.strip() == "can prospect"
            ].copy()
            filtered_out = before - len(prospectable)
        else:
            prospectable = all_companies.copy()
            filtered_out = 0

        # Find other useful columns
        aov_col = None
        hc_col = None
        website_col = None
        for col in all_companies.columns:
            cl = _col_lower(col)
            if "aov" in cl:
                aov_col = col
            if "headcount" in cl or "head count" in cl or "employee" in cl:
                hc_col = col
            if "website" in cl or "url" in cl:
                website_col = col

        # Display summary
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Accounts", len(all_companies))
        col2.metric("Can Prospect", len(prospectable))
        col3.metric("Filtered Out", filtered_out)
        col4.metric("Team Members", prospectable["_team_member"].nunique())

        # Show preview
        with st.expander("📋 Preview Accounts", expanded=False):
            display_cols = [name_col, "_team_member"]
            if status_col:
                display_cols.append(status_col)
            if aov_col:
                display_cols.append(aov_col)
            if hc_col:
                display_cols.append(hc_col)
            st.dataframe(
                prospectable[display_cols].head(20),
                width="stretch",
            )

        st.divider()

        # ── Run Analysis ────────────────────────────────────────────
        if st.button("🚀 Analyze & Score Accounts", type="primary", width="stretch"):
            if not exa_key:
                st.error("Please enter your Exa API key in the sidebar.")
                st.stop()

            from enrichment import enrich_company, score_company

            progress_bar = st.progress(0, text="Starting enrichment...")
            status_text = st.empty()
            log_expander = st.expander("📝 Live Log", expanded=True)
            log_area = log_expander.empty()
            log_lines = []

            def add_log(msg):
                log_lines.append(f"`{datetime.now().strftime('%H:%M:%S')}` {msg}")
                log_area.markdown("\n\n".join(log_lines[-20:]))  # show last 20

            add_log(f"🚀 Starting enrichment of {len(prospectable)} companies...")
            add_log(f"Exa API key: {'✅ configured' if exa_key else '❌ MISSING'}")
            add_log(f"Layers: Website={do_website}, Tech={do_tech_stack}, Intel={do_company_intel}, News={do_news}, Execs={do_executives}")

            results = []
            total = len(prospectable)

            for idx, (_, row) in enumerate(prospectable.iterrows()):
                company_name = row[name_col]
                # Skip empty/nan company names
                if pd.isna(company_name) or str(company_name).strip() == "":
                    add_log(f"[{idx+1}/{total}] ⏭️ Skipping empty row")
                    continue

                company_name = str(company_name).strip()
                team = row["_team_member"]
                aov = row[aov_col] if aov_col and pd.notna(row.get(aov_col, None)) else ""
                headcount = row[hc_col] if hc_col and pd.notna(row.get(hc_col, None)) else ""
                website = row[website_col] if website_col and pd.notna(row.get(website_col)) else None

                pct = (idx + 1) / total
                progress_bar.progress(pct, text=f"[{idx+1}/{total}] {company_name}")
                status_text.text(f"Enriching: {company_name} ({team})")

                try:
                    enriched = enrich_company(
                        company_name=company_name,
                        team_member=team,
                        aov_band=aov,
                        headcount=headcount,
                        known_website=website,
                        exa_key=exa_key,
                        do_website=do_website,
                        do_tech_stack=do_tech_stack,
                        do_company_intel=do_company_intel,
                        do_news=do_news,
                        do_executives=do_executives,
                        do_job_postings=do_job_postings,
                    )
                    # Log what we found
                    parts = []
                    if enriched.get("website"):
                        parts.append(f"🌐 {enriched['website']}")
                    if enriched.get("technologies"):
                        parts.append(f"🔧 {len(enriched['technologies'])} techs")
                    if enriched.get("industry"):
                        parts.append(f"🏢 {enriched['industry']}")
                    if enriched.get("news_items"):
                        parts.append(f"📰 {len(enriched['news_items'])} news")
                    if enriched.get("executive_changes"):
                        parts.append(f"👔 {len(enriched['executive_changes'])} execs")
                    summary = " | ".join(parts) if parts else "⚠️ no data found"
                    add_log(f"[{idx+1}/{total}] **{company_name}** ({team}) — {summary}")

                except Exception as e:
                    enriched = {
                        "name": company_name,
                        "team_member": team,
                        "error": str(e),
                    }
                    add_log(f"[{idx+1}/{total}] **{company_name}** ({team}) — ❌ Error: {e}")

                # Score
                weights = {
                    "tech_stack": w_tech,
                    "compelling_events": w_events,
                    "new_executives": w_execs,
                    "company_size": w_size,
                    "accessibility": w_access,
                }
                enriched = score_company(enriched, weights)
                results.append(enriched)

            progress_bar.progress(1.0, text="✅ Complete!")
            scored_count = sum(1 for r in results if r.get("score", 0) > 20)
            add_log(f"✅ Done! {scored_count}/{total} companies enriched with meaningful data")
            status_text.text(f"Enriched {total} companies — {scored_count} with meaningful scores")

            st.session_state.results = results

    except Exception as e:
        st.error(f"Error reading file: {e}")
        import traceback
        st.code(traceback.format_exc())

# ── Display Results ─────────────────────────────────────────────────
if st.session_state.results:
    results = st.session_state.results
    st.divider()
    st.header(f"📊 Results ({len(results)} accounts)")

    # Build results DataFrame
    rows = []
    for r in results:
        # Build news with links
        news_parts = []
        for n in r.get("news_with_links", r.get("news_items", []))[:3]:
            if isinstance(n, dict):
                events = n.get("events", [])
                tag = f"[{', '.join(events)}] " if events else ""
                title = n.get("text", n.get("title", ""))
                url = n.get("url", "")
                if url:
                    news_parts.append(f"{tag}{title} ({url})")
                else:
                    news_parts.append(f"{tag}{title}")
        news_str = "\n".join(news_parts) if news_parts else ""

        # Build executives with links
        exec_parts = []
        for e in r.get("executives_with_links", r.get("executive_changes", []))[:3]:
            if isinstance(e, dict):
                name = e.get("text", e.get("person_name", "")) or e.get("title", "")
                li = e.get("linkedin_url", "")
                ann = e.get("announcement_url", "")
                links = []
                if li:
                    links.append(f"LinkedIn: {li}")
                if ann:
                    links.append(f"Article: {ann}")
                link_str = f" | {' | '.join(links)}" if links else ""
                exec_parts.append(f"{name}{link_str}")
        exec_str = "\n".join(exec_parts) if exec_parts else ""

        row = {
            "Rank": 0,
            "Score": r.get("score", 0),
            "Account": r.get("name", ""),
            "AOV Band": r.get("aov_band", ""),
            "Team Member": r.get("team_member", ""),
            "Website": r.get("website", ""),
            "Industry": r.get("industry", ""),
            "Description": r.get("description", ""),
            "Headcount": str(r.get("headcount", "")) if r.get("headcount") else "",
            "Tech Stack Score": r.get("tech_score", 0),
            "Key Technologies": r.get("key_technologies", ""),
            "CRM": r.get("crm", ""),
            "SF Opportunity": r.get("sf_opportunity", ""),
            "Recent News": news_str,
            "Compelling Events": ", ".join(r.get("compelling_events", [])) if isinstance(r.get("compelling_events"), list) else str(r.get("compelling_events", "")),
            "New Executives": exec_str,
            "Top Signals": "\n".join(
                re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', sig)  # strip markdown links for table display
                for sig in r.get("top_signals", [])[:3]
            ),
            "Recommended Contacts": "\n".join(
                re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', c)
                for c in r.get("recommended_contacts", [])[:4]
            ),
        }
        rows.append(row)

    df_results = pd.DataFrame(rows)
    df_results = df_results.sort_values("Score", ascending=False).reset_index(drop=True)
    df_results["Rank"] = range(1, len(df_results) + 1)

    # Summary metrics
    st.markdown("---")
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    sf_existing = sum(1 for r in results if "expansion" in r.get("sf_opportunity", "").lower())
    sf_greenfield = sum(1 for r in results if "greenfield" in r.get("sf_opportunity", "").lower())
    sf_displacement = sum(1 for r in results if "displacement" in r.get("sf_opportunity", "").lower())
    jp_enriched = sum(1 for r in results if r.get("job_posting_tools"))
    col1.metric("Accounts Scored", len(results))
    col2.metric("Expansion", sf_existing)
    col3.metric("Displacement", sf_displacement)
    col4.metric("Greenfield", sf_greenfield)
    col5.metric("Job Post Intel", jp_enriched)
    col6.metric("Avg Score", f"{df_results['Score'].mean():.0f}")

    # Filter by team member
    team_filter = st.multiselect(
        "Filter by Team Member",
        options=sorted(df_results["Team Member"].unique()),
        default=sorted(df_results["Team Member"].unique()),
    )
    filtered = df_results[df_results["Team Member"].isin(team_filter)]

    # Reorder columns so reps see the most actionable info first
    priority_cols = [
        "Rank", "Score", "Account", "AOV Band", "SF Opportunity", "Team Member",
        "Top Signals", "CRM", "Recommended Contacts",
        "Industry", "Headcount", "Website",
        "Key Technologies", "Tech Stack Score",
        "Compelling Events", "Recent News", "New Executives",
        "Description",
    ]
    display_cols_ordered = [c for c in priority_cols if c in filtered.columns]
    # Add any remaining columns not in priority list
    for c in filtered.columns:
        if c not in display_cols_ordered:
            display_cols_ordered.append(c)
    filtered = filtered[display_cols_ordered]

    # Display table
    st.dataframe(
        filtered,
        width="stretch",
        height=600,
        column_config={
            "Rank": st.column_config.NumberColumn("Rank", width="small"),
            "Score": st.column_config.ProgressColumn(
                "Score", min_value=0, max_value=100, format="%d", width="small"
            ),
            "Account": st.column_config.TextColumn("Account", width="medium"),
            "AOV Band": st.column_config.TextColumn("AOV Band", width="small"),
            "SF Opportunity": st.column_config.TextColumn("SF Opp", width="small"),
            "Top Signals": st.column_config.TextColumn("Top Signals", width="large"),
            "CRM": st.column_config.TextColumn("CRM", width="small"),
            "Recommended Contacts": st.column_config.TextColumn("Who to Reach", width="large"),
            "Tech Stack Score": st.column_config.ProgressColumn(
                "Tech Score", min_value=0, max_value=100, format="%d", width="small"
            ),
            "Website": st.column_config.LinkColumn("Website", width="medium"),
        },
    )

    # Download buttons
    st.divider()
    col1, col2, col3, col4 = st.columns(4)

    # Excel download with hyperlinks
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def build_excel(results_list):
        wb = Workbook()
        ws = wb.active
        ws.title = "Account Scores"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        link_font = Font(color="0563C1", underline="single")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        headers = [
            "Rank", "Score", "Account", "AOV Band", "SF Opportunity",
            "Team Member", "Top Signals", "Recommended Contacts",
            "CRM", "Industry", "Headcount", "Website",
            "Tech Score", "Key Technologies",
            "Compelling Events", "Recent News", "New Executives",
            "Job Posting Tech", "Description", "News Links", "LinkedIn Links",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        sorted_results = sorted(results_list, key=lambda x: -x.get("score", 0))

        for idx, r in enumerate(sorted_results, 2):
            rank = idx - 1
            # New column order: Rank(1), Score(2), Account(3), AOV Band(4),
            # SF Opp(5), Team(6), Top Signals(7), Recommended Contacts(8),
            # CRM(9), Industry(10), Headcount(11), Website(12),
            # Tech Score(13), Key Tech(14), Events(15), News(16),
            # Execs(17), JP Tech(18), Description(19), News Links(20), LI Links(21)

            ws.cell(row=idx, column=1, value=rank).border = thin_border
            ws.cell(row=idx, column=2, value=r.get("score", 0)).border = thin_border
            ws.cell(row=idx, column=3, value=r.get("name", "")).border = thin_border

            # AOV Band (col 4)
            aov_cell = ws.cell(row=idx, column=4, value=str(r.get("aov_band", "")))
            aov_cell.border = thin_border
            if r.get("aov_band"):
                aov_cell.fill = highlight_fill

            # SF Opportunity (col 5)
            sf_cell = ws.cell(row=idx, column=5, value=r.get("sf_opportunity", ""))
            sf_cell.border = thin_border
            if r.get("sf_opportunity") in ("Expansion", "Displacement"):
                sf_cell.fill = highlight_fill

            # Team Member (col 6)
            ws.cell(row=idx, column=6, value=r.get("team_member", "")).border = thin_border

            # Top Signals (col 7) — THE most important column for reps
            top_sigs = r.get("top_signals", [])
            sig_cell = ws.cell(row=idx, column=7, value="\n".join(top_sigs[:3]))
            sig_cell.border = thin_border
            sig_cell.alignment = Alignment(wrap_text=True, vertical='top')
            if top_sigs:
                sig_cell.fill = highlight_fill

            # Recommended Contacts (col 8)
            rec_contacts = r.get("recommended_contacts", [])
            rc_cell = ws.cell(row=idx, column=8, value="\n".join(rec_contacts[:4]))
            rc_cell.border = thin_border
            rc_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # CRM (col 9)
            ws.cell(row=idx, column=9, value=r.get("crm", "")).border = thin_border

            # Industry (col 10)
            ws.cell(row=idx, column=10, value=r.get("industry", "")).border = thin_border

            # Headcount (col 11)
            hc = r.get("headcount", "")
            ws.cell(row=idx, column=11, value=str(hc) if hc else "").border = thin_border

            # Website as hyperlink (col 12)
            url = r.get("website", "")
            if url:
                cell = ws.cell(row=idx, column=12)
                cell.value = url
                cell.hyperlink = url
                cell.font = link_font
                cell.border = thin_border
            else:
                ws.cell(row=idx, column=12, value="").border = thin_border

            # Tech Score (col 13)
            ws.cell(row=idx, column=13, value=r.get("tech_score", 0)).border = thin_border

            # Key Technologies (col 14)
            ws.cell(row=idx, column=14, value=r.get("key_technologies", "")).border = thin_border

            # Compelling events (col 15)
            events = r.get("compelling_events", [])
            events_str = ", ".join(events) if isinstance(events, list) else str(events)
            ev_cell = ws.cell(row=idx, column=15, value=events_str)
            ev_cell.border = thin_border
            ev_cell.alignment = Alignment(wrap_text=True)
            if events:
                ev_cell.fill = highlight_fill

            # News (col 16)
            news_items = r.get("news_items", [])
            if news_items:
                first_news = news_items[0]
                first_title = first_news.get("title", "")
                first_url = first_news.get("url", "")
                events_tag = ""
                if first_news.get("events"):
                    events_tag = f"[{', '.join(first_news['events'])}] "
                extra_lines = []
                for n in news_items[1:5]:
                    et = ""
                    if n.get("events"):
                        et = f"[{', '.join(n['events'])}] "
                    extra_lines.append(f"{et}{n.get('title', '')}")
                full_text = f"{events_tag}{first_title}"
                if extra_lines:
                    full_text += "\n" + "\n".join(extra_lines)
                news_cell = ws.cell(row=idx, column=16, value=full_text)
                if first_url:
                    news_cell.hyperlink = first_url
                    news_cell.font = link_font
            else:
                news_cell = ws.cell(row=idx, column=16, value="")
            news_cell.border = thin_border
            news_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Executives (col 17)
            exec_changes = r.get("executive_changes", [])
            if exec_changes:
                first_exec = exec_changes[0]
                person = first_exec.get("person_name", "") or "New hire"
                etitle = first_exec.get("title", "")
                li = first_exec.get("linkedin_url", "")
                full_text = f"{person} — {etitle}"
                extra_lines = []
                for e in exec_changes[1:3]:
                    p = e.get("person_name", "") or "New hire"
                    t = e.get("title", "")
                    extra_lines.append(f"{p} — {t}")
                if extra_lines:
                    full_text += "\n" + "\n".join(extra_lines)
                exec_cell = ws.cell(row=idx, column=17, value=full_text)
                if li:
                    exec_cell.hyperlink = li
                    exec_cell.font = link_font
            else:
                exec_cell = ws.cell(row=idx, column=17, value="")
            exec_cell.border = thin_border
            exec_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Job Posting Tech (col 18)
            jp_tools = r.get("job_posting_tools", [])
            if jp_tools:
                jp_lines = [f"{t['name']} ({t['category']})" for t in jp_tools]
                jp_cell = ws.cell(row=idx, column=18, value="\n".join(jp_lines))
                jp_cell.fill = highlight_fill
            else:
                jp_cell = ws.cell(row=idx, column=18, value="")
            jp_cell.border = thin_border
            jp_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Description (col 19)
            desc_cell = ws.cell(row=idx, column=19, value=r.get("description", "")[:500])
            desc_cell.border = thin_border
            desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # News Links (col 20)
            news_urls = []
            for n in (r.get("news_items", []))[:5]:
                url = n.get("url", "")
                if url:
                    news_urls.append(url)
            if news_urls:
                nl_cell = ws.cell(row=idx, column=20, value="\n".join(news_urls))
                nl_cell.hyperlink = news_urls[0]
                nl_cell.font = link_font
            else:
                nl_cell = ws.cell(row=idx, column=20, value="")
            nl_cell.border = thin_border
            nl_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # LinkedIn Links (col 21)
            li_urls = []
            for e in (r.get("executive_changes", []))[:5]:
                li = e.get("linkedin_url", "")
                if li:
                    li_urls.append(li)
            if li_urls:
                li_cell = ws.cell(row=idx, column=21, value="\n".join(li_urls))
                li_cell.hyperlink = li_urls[0]
                li_cell.font = link_font
            else:
                li_cell = ws.cell(row=idx, column=21, value="")
            li_cell.border = thin_border
            li_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Column widths — prioritize action columns (signals, contacts)
        widths = {
            'A': 6, 'B': 8, 'C': 28, 'D': 14, 'E': 16,
            'F': 16, 'G': 50, 'H': 45, 'I': 15,
            'J': 18, 'K': 12, 'L': 30, 'M': 10,
            'N': 25, 'O': 25, 'P': 50, 'Q': 45,
            'R': 35, 'S': 50, 'T': 40, 'U': 40,
        }
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:U1"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    xlsx_data = build_excel(results)
    col1.download_button(
        "📥 Download Excel (with links)",
        data=xlsx_data,
        file_name=f"account_scores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # CSV download
    csv_data = filtered.to_csv(index=False)
    col2.download_button(
        "📥 Download CSV",
        data=csv_data,
        file_name=f"account_scores_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
    )

    # HTML shareable report
    def build_html_report(results_list):
        sorted_results = sorted(results_list, key=lambda x: -x.get("score", 0))
        sf_exp = sum(1 for r in sorted_results if "expansion" in r.get("sf_opportunity", "").lower())
        sf_dis = sum(1 for r in sorted_results if "displacement" in r.get("sf_opportunity", "").lower())
        sf_grn = sum(1 for r in sorted_results if "greenfield" in r.get("sf_opportunity", "").lower())
        avg_sc = round(sum(r.get("score", 0) for r in sorted_results) / max(len(sorted_results), 1))
        gen_date = datetime.now().strftime("%B %d, %Y at %I:%M %p")

        # Group by team member for rep sheets
        team_groups = {}
        for r in sorted_results:
            tm = r.get("team_member", "Unknown")
            team_groups.setdefault(tm, []).append(r)

        def _score_color(s):
            if s >= 70: return "#15803D"
            if s >= 40: return "#B45309"
            return "#DC2626"

        def _opp_color(opp):
            return {"Expansion": "#15803D", "Displacement": "#B45309", "Greenfield": "#0176D3"}.get(opp, "#64748B")

        def _esc(text):
            if not text: return ""
            return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

        def _linkify(text):
            """Convert markdown links [text](url) to HTML <a> tags."""
            if not text: return ""
            return re.sub(
                r'\[([^\]]+)\]\(([^)]+)\)',
                r'<a href="\2" target="_blank" style="color:#0176D3;">\1</a>',
                _esc(text)
            )

        # Build ranked table rows
        table_rows = ""
        for rank, r in enumerate(sorted_results, 1):
            score = r.get("score", 0)
            sc = _score_color(score)
            opp = r.get("sf_opportunity", "")
            oc = _opp_color(opp)
            signals = "<br>".join(_linkify(s) for s in r.get("top_signals", [])[:3])
            contacts = "<br>".join(_linkify(c) for c in r.get("recommended_contacts", [])[:3])
            website = r.get("website", "")
            website_link = f'<a href="{_esc(website)}" target="_blank" style="color:#0176D3;">{_esc(website)}</a>' if website else ""

            table_rows += f"""
            <tr>
                <td style="text-align:center;font-weight:700;">{rank}</td>
                <td style="text-align:center;">
                    <span style="background:{sc};color:white;padding:3px 10px;border-radius:6px;font-weight:700;font-size:0.85rem;">{score}</span>
                </td>
                <td style="font-weight:600;">{_esc(r.get('name', ''))}</td>
                <td>{_esc(r.get('aov_band', ''))}</td>
                <td><span style="background:{oc}15;color:{oc};padding:2px 8px;border-radius:4px;font-weight:600;font-size:0.8rem;">{_esc(opp)}</span></td>
                <td>{_esc(r.get('team_member', ''))}</td>
                <td style="font-size:0.85rem;">{signals}</td>
                <td>{_esc(r.get('crm', ''))}</td>
                <td style="font-size:0.85rem;">{contacts}</td>
                <td>{website_link}</td>
            </tr>"""

        # Build rep action sheets
        rep_sections = ""
        for tm in sorted(team_groups.keys()):
            accts = team_groups[tm][:5]
            cards = ""
            for i, acct in enumerate(accts, 1):
                score = acct.get("score", 0)
                sc = _score_color(score)
                opp = acct.get("sf_opportunity", "")
                oc = _opp_color(opp)
                aov = acct.get("aov_band", "")
                crm_val = acct.get("crm", "")
                website = acct.get("website", "")

                sig_items = "".join(f"<li>{_linkify(s)}</li>" for s in acct.get("top_signals", [])[:3])
                contact_items = "".join(f"<li>{_linkify(c)}</li>" for c in acct.get("recommended_contacts", [])[:3])

                aov_html = f"<span style='color:#64748B;font-size:0.85rem;'>AOV: {_esc(aov)}</span>" if aov else ""
                website_html = (f' · <a href="{_esc(website)}" target="_blank" style="color:#0176D3;">{_esc(website)}</a>' if website else "")
                sig_html = (f'<div style="margin-bottom:0.5rem;"><strong style="font-size:0.9rem;">Why call:</strong>'
                            f'<ul style="margin:0.25rem 0 0 1.25rem;color:#334155;font-size:0.85rem;">{sig_items}</ul></div>' if sig_items else "")
                contact_html = (f'<div><strong style="font-size:0.9rem;">Who to reach:</strong>'
                                f'<ul style="margin:0.25rem 0 0 1.25rem;color:#334155;font-size:0.85rem;">{contact_items}</ul></div>' if contact_items else "")

                cards += f"""
                <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:12px;padding:1rem 1.25rem;margin-bottom:0.75rem;">
                    <div style="display:flex;align-items:center;gap:0.75rem;margin-bottom:0.5rem;flex-wrap:wrap;">
                        <span style="background:{sc};color:white;font-weight:700;padding:4px 10px;border-radius:6px;font-size:0.85rem;">#{i} — {score}/100</span>
                        <span style="font-weight:700;font-size:1.05rem;color:#0F172A;">{_esc(acct.get('name', ''))}</span>
                        <span style="background:{oc}15;color:{oc};font-weight:600;padding:3px 8px;border-radius:4px;font-size:0.8rem;">{_esc(opp)}</span>
                        {aov_html}
                    </div>
                    <div style="color:#475569;font-size:0.85rem;margin-bottom:0.5rem;">
                        CRM: <strong>{_esc(crm_val)}</strong>{website_html}
                    </div>
                    {sig_html}
                    {contact_html}
                </div>"""

            rep_sections += f"""
            <div style="margin-bottom:2rem;">
                <div onclick="this.nextElementSibling.style.display=this.nextElementSibling.style.display==='none'?'block':'none';this.querySelector('.arrow').textContent=this.nextElementSibling.style.display==='none'?'▸':'▾'"
                     style="cursor:pointer;background:#F1F5F9;border-radius:10px;padding:0.75rem 1rem;font-weight:700;font-size:1.05rem;color:#0F172A;display:flex;align-items:center;gap:0.5rem;">
                    <span class="arrow">▸</span> {_esc(tm)} — Top {len(accts)} accounts
                </div>
                <div style="display:none;padding-top:0.75rem;">
                    {cards}
                </div>
            </div>"""

        # Build account deep dive sections
        deep_dive_sections = ""
        for r in sorted_results:
            name = _esc(r.get("name", ""))
            score = r.get("score", 0)
            sc = _score_color(score)

            # Tech stack table
            tech_rows = ""
            sf_relevant = [
                t for t in r.get("technologies", [])
                if t.get("category") not in HIDE_CATEGORIES
            ]
            jp_tools = r.get("job_posting_tools", [])
            for t in jp_tools:
                if t["name"].lower() not in [x.get("name", "").lower() for x in sf_relevant]:
                    sf_relevant.append({"name": t["name"], "category": t.get("category", ""), "source": "Job Posting"})

            for t in sf_relevant[:15]:
                tname = t.get("name", "")
                cat = t.get("category", "")
                sf_sol = SALESFORCE_SOLUTION_MAP.get(tname, "")
                if not sf_sol:
                    sf_sol = _get_sf_solution(tname, cat)
                # Skip tools with no Salesforce compete story
                if sf_sol == "—" or not sf_sol:
                    continue
                source = t.get("source", "Website")
                tech_rows += f"<tr><td>{_esc(cat)}</td><td>{_esc(tname)}</td><td>{_esc(sf_sol)}</td><td>{_esc(source)}</td></tr>"

            # News items
            news_rows = ""
            for n in r.get("news_with_links", r.get("news_items", []))[:5]:
                if isinstance(n, dict):
                    events = n.get("events", [])
                    tag = f"[{', '.join(events)}] " if events else ""
                    title = n.get("text", n.get("title", ""))
                    url = n.get("url", "")
                    date = n.get("date", "")
                    link = f'<a href="{_esc(url)}" target="_blank" style="color:#0176D3;">{_esc(title)}</a>' if url else _esc(title)
                    news_rows += f"<tr><td>{_esc(tag)}</td><td>{link}</td><td>{_esc(date)}</td></tr>"

            # Exec changes
            exec_rows = ""
            for e in r.get("executives_with_links", r.get("executive_changes", []))[:5]:
                if isinstance(e, dict):
                    pname = _esc(e.get("person_name", "") or e.get("text", ""))
                    title = _esc(e.get("title", e.get("matched_title", "")))
                    li = e.get("linkedin_url", "")
                    li_link = f'<a href="{_esc(li)}" target="_blank" style="color:#0176D3;">LinkedIn</a>' if li else ""
                    exec_rows += f"<tr><td>{pname}</td><td>{title}</td><td>{li_link}</td></tr>"

            _tbl = 'style="width:100%;border-collapse:collapse;font-size:0.85rem;"'
            _th = "style='text-align:left;padding:6px 10px;'"
            _dd_tech_html = (
                f"<h4 style='color:#0F172A;margin:1rem 0 0.5rem;'>Tech Stack</h4>"
                f"<table {_tbl}><tr style='background:#F1F5F9;'><th {_th}>Category</th><th {_th}>Current Tool</th><th {_th}>SF Solution</th><th {_th}>Source</th></tr>"
                f"{tech_rows}</table>"
            ) if tech_rows else ""
            _dd_news_html = (
                f"<h4 style='color:#0F172A;margin:1rem 0 0.5rem;'>Recent News</h4>"
                f"<table {_tbl}><tr style='background:#F1F5F9;'><th {_th}>Signal</th><th {_th}>Headline</th><th {_th}>Date</th></tr>"
                f"{news_rows}</table>"
            ) if news_rows else ""
            _dd_exec_html = (
                f"<h4 style='color:#0F172A;margin:1rem 0 0.5rem;'>Executive Changes</h4>"
                f"<table {_tbl}><tr style='background:#F1F5F9;'><th {_th}>Name</th><th {_th}>Title</th><th {_th}>LinkedIn</th></tr>"
                f"{exec_rows}</table>"
            ) if exec_rows else ""

            deep_dive_sections += f"""
            <div style="margin-bottom:1.5rem;">
                <div onclick="this.nextElementSibling.style.display=this.nextElementSibling.style.display==='none'?'block':'none';this.querySelector('.arrow').textContent=this.nextElementSibling.style.display==='none'?'▸':'▾'"
                     style="cursor:pointer;background:#F1F5F9;border-radius:10px;padding:0.75rem 1rem;font-weight:600;font-size:1rem;color:#0F172A;display:flex;align-items:center;gap:0.5rem;">
                    <span class="arrow">▸</span>
                    <span style="background:{sc};color:white;padding:2px 8px;border-radius:4px;font-size:0.8rem;font-weight:700;">{score}</span>
                    {name} — {_esc(r.get('sf_opportunity', ''))}
                </div>
                <div style="display:none;padding:1rem 0;">
                    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:1rem;margin-bottom:1rem;">
                        <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;padding:0.75rem;text-align:center;">
                            <div style="color:#64748B;font-size:0.75rem;text-transform:uppercase;">CRM</div>
                            <div style="font-weight:700;color:#0F172A;">{_esc(r.get('crm', 'Unknown'))}</div>
                        </div>
                        <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;padding:0.75rem;text-align:center;">
                            <div style="color:#64748B;font-size:0.75rem;text-transform:uppercase;">Industry</div>
                            <div style="font-weight:700;color:#0F172A;">{_esc(r.get('industry', '—'))}</div>
                        </div>
                        <div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;padding:0.75rem;text-align:center;">
                            <div style="color:#64748B;font-size:0.75rem;text-transform:uppercase;">Headcount</div>
                            <div style="font-weight:700;color:#0F172A;">{_esc(str(r.get('headcount', '—')))}</div>
                        </div>
                    </div>
                    {_dd_tech_html}
                    {_dd_news_html}
                    {_dd_exec_html}
                </div>
            </div>"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Account Scorer Report — {gen_date}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; color: #0F172A; background: #F8FAFC; }}
.header {{ background: linear-gradient(135deg, #032D60 0%, #0176D3 100%); color: white; padding: 2rem 3rem; }}
.header h1 {{ font-size: 1.8rem; font-weight: 700; letter-spacing: -0.03em; }}
.header p {{ opacity: 0.8; margin-top: 0.25rem; font-size: 0.95rem; }}
.container {{ max-width: 1400px; margin: 0 auto; padding: 2rem; }}
.metrics {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 1rem; margin-bottom: 2rem; }}
.metric {{ background: white; border: 1px solid #E2E8F0; border-radius: 12px; padding: 1rem 1.25rem; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }}
.metric .label {{ color: #64748B; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.05em; }}
.metric .value {{ font-size: 1.5rem; font-weight: 700; color: #0F172A; }}
h2 {{ font-size: 1.4rem; font-weight: 700; color: #0F172A; letter-spacing: -0.02em; margin: 2rem 0 1rem; }}
table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
th {{ background: #F1F5F9; text-align: left; padding: 10px 12px; font-weight: 600; font-size: 0.8rem; color: #475569; text-transform: uppercase; letter-spacing: 0.03em; }}
td {{ padding: 10px 12px; border-top: 1px solid #F1F5F9; vertical-align: top; font-size: 0.88rem; }}
tr:hover {{ background: #F8FAFC; }}
a {{ color: #0176D3; text-decoration: none; }}
a:hover {{ text-decoration: underline; }}
.section {{ margin-bottom: 2rem; }}
.legend {{ background: white; border: 1px solid #E2E8F0; border-radius: 8px; padding: 0.75rem 1rem; font-size: 0.8rem; color: #64748B; margin-bottom: 1.5rem; }}
.footer {{ text-align: center; color: #94A3B8; font-size: 0.8rem; padding: 2rem; border-top: 1px solid #E2E8F0; margin-top: 3rem; }}
@media print {{ .header {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }} }}
@media (max-width: 768px) {{ .metrics {{ grid-template-columns: repeat(2, 1fr); }} table {{ font-size: 0.8rem; }} }}
</style>
</head>
<body>
<div class="header">
    <h1>☁️ Account Scorer Report</h1>
    <p>Generated {gen_date} · {len(sorted_results)} accounts scored</p>
</div>
<div class="container">
    <div class="metrics">
        <div class="metric"><div class="label">Accounts Scored</div><div class="value">{len(sorted_results)}</div></div>
        <div class="metric"><div class="label">Expansion</div><div class="value" style="color:#15803D;">{sf_exp}</div></div>
        <div class="metric"><div class="label">Displacement</div><div class="value" style="color:#B45309;">{sf_dis}</div></div>
        <div class="metric"><div class="label">Greenfield</div><div class="value" style="color:#0176D3;">{sf_grn}</div></div>
        <div class="metric"><div class="label">Avg Score</div><div class="value">{avg_sc}</div></div>
    </div>

    <div class="legend">
        Signal Legend: 🔴 M&amp;A · 🟢 Funding/IPO · 🟡 Contract/Deal · 🔵 Partnership · 🟣 Expansion · 🟠 Digital Transformation · ⚪ Restructuring
    </div>

    <div class="section">
        <h2>📊 Ranked Accounts</h2>
        <table>
            <tr>
                <th>#</th><th>Score</th><th>Account</th><th>AOV</th><th>Opportunity</th>
                <th>Rep</th><th>Top Signals</th><th>CRM</th><th>Contacts</th><th>Website</th>
            </tr>
            {table_rows}
        </table>
    </div>

    <div class="section">
        <h2>📋 Rep Action Sheets</h2>
        <p style="color:#64748B;margin-bottom:1rem;font-size:0.9rem;">Click a rep name to expand their top accounts.</p>
        {rep_sections}
    </div>

    <div class="section">
        <h2>🔍 Account Deep Dives</h2>
        <p style="color:#64748B;margin-bottom:1rem;font-size:0.9rem;">Click an account to see tech stack, news, and executive details.</p>
        {deep_dive_sections}
    </div>

    <div class="footer">
        Account Scorer — Salesforce Account Prioritization Tool · Generated {gen_date}
    </div>
</div>
</body>
</html>"""
        return html

    html_report = build_html_report(results)
    col3.download_button(
        "🌐 Download Shareable Report",
        data=html_report,
        file_name=f"account_report_{datetime.now().strftime('%Y%m%d_%H%M')}.html",
        mime="text/html",
    )

    # JSON download
    json_data = json.dumps(results, indent=2, default=str)
    col4.download_button(
        "📥 Download Full JSON",
        data=json_data,
        file_name=f"account_enrichment_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
        mime="application/json",
    )

    # ── Rep Action Sheets ─────────────────────────────────────────────
    st.divider()
    st.markdown("""
    <h2 style="font-size: 1.5rem; font-weight: 700; color: #0F172A; letter-spacing: -0.02em;">
        📋 Rep Action Sheets
    </h2>
    <p style="color: #64748B; margin-top: 0.25rem;">
        Top 5 accounts per rep with clear reasons to call and people to reach.
        Hand these to your reps for immediate action.
    </p>
    """, unsafe_allow_html=True)

    # Group results by team member, sorted by score within each group
    team_groups = {}
    for r in sorted(results, key=lambda x: -x.get("score", 0)):
        tm = r.get("team_member", "Unknown")
        if tm not in team_groups:
            team_groups[tm] = []
        team_groups[tm].append(r)

    rep_top_n = st.slider("Accounts per rep", 3, 10, 5, key="rep_top_n")

    for team_member in sorted(team_groups.keys()):
        accts = team_groups[team_member][:rep_top_n]
        with st.expander(f"**{team_member}** — Top {len(accts)} accounts", expanded=False):
            for i, acct in enumerate(accts, 1):
                score = acct.get("score", 0)
                name = acct.get("name", "")
                sf_opp = acct.get("sf_opportunity", "Unknown")
                aov = acct.get("aov_band", "")
                crm_val = acct.get("crm", "")
                website = acct.get("website", "")

                # Score color
                if score >= 70:
                    score_color = "#15803D"
                elif score >= 40:
                    score_color = "#B45309"
                else:
                    score_color = "#DC2626"

                # SF Opportunity badge
                opp_colors = {"Expansion": "#15803D", "Displacement": "#B45309", "Greenfield": "#0176D3"}
                opp_color = opp_colors.get(sf_opp, "#64748B")

                st.markdown(f"""
                <div style="background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 12px;
                            padding: 1rem 1.25rem; margin-bottom: 0.75rem;">
                    <div style="display: flex; align-items: center; gap: 1rem; margin-bottom: 0.5rem;">
                        <span style="background: {score_color}; color: white; font-weight: 700;
                                     padding: 4px 10px; border-radius: 6px; font-size: 0.9rem;">
                            #{i} — {score}/100
                        </span>
                        <span style="font-weight: 700; font-size: 1.1rem; color: #0F172A;">{name}</span>
                        <span style="background: {opp_color}15; color: {opp_color}; font-weight: 600;
                                     padding: 3px 8px; border-radius: 4px; font-size: 0.8rem;">
                            {sf_opp}
                        </span>
                        {"<span style='color: #64748B; font-size: 0.85rem;'>AOV: " + str(aov) + "</span>" if aov else ""}
                    </div>
                    <div style="color: #475569; font-size: 0.85rem; margin-bottom: 0.25rem;">
                        CRM: <strong>{crm_val}</strong>
                        {" · <a href='" + website + "' target='_blank'>" + website + "</a>" if website else ""}
                    </div>
                </div>
                """, unsafe_allow_html=True)

                # Top signals as bullet points
                top_signals = acct.get("top_signals", [])
                if top_signals:
                    st.markdown("**Why call:**")
                    for sig in top_signals[:3]:
                        st.markdown(f"  - {sig}")

                # Recommended contacts
                rec_contacts = acct.get("recommended_contacts", [])
                if rec_contacts:
                    st.markdown("**Who to reach:**")
                    for c in rec_contacts[:3]:
                        st.markdown(f"  - {c}")

                st.markdown("---")

    # ── Per-Account Detail View ─────────────────────────────────────
    st.divider()
    st.markdown("""
    <h2 style="font-size: 1.5rem; font-weight: 700; color: #0F172A; letter-spacing: -0.02em;">
        Account Deep Dive
    </h2>
    """, unsafe_allow_html=True)

    account_names = [r.get("name", "") for r in sorted(results, key=lambda x: -x.get("score", 0))]
    selected = st.selectbox("Select Account", account_names)

    if selected:
        acct = next(r for r in results if r.get("name") == selected)

        col1, col2, col3 = st.columns(3)
        col1.metric("Overall Score", f"{acct.get('score', 0)}/100")
        col2.metric("Tech Score", f"{acct.get('tech_score', 0)}/100")
        col3.metric("SF Opportunity", acct.get("sf_opportunity", "Unknown"))

        # ── Top Signals & Recommended Contacts ─────────────────
        sig_col, contact_col = st.columns(2)
        with sig_col:
            st.markdown("#### 🎯 Top Signals — Why Reach Out")
            top_signals = acct.get("top_signals", [])
            if top_signals:
                for sig in top_signals:
                    st.markdown(f"- {sig}")
            else:
                st.caption("No strong signals detected")

        with contact_col:
            st.markdown("#### 👥 Who to Reach")
            contacts = acct.get("recommended_contacts", [])
            if contacts:
                for c in contacts[:6]:
                    st.markdown(f"- {c}")
            else:
                st.caption("No specific contacts identified")

        st.divider()
        tab1, tab2, tab3, tab4 = st.tabs(["🔧 Tech Stack", "📰 News", "👔 Executives", "📋 Full Data"])

        with tab1:
            sf_relevant = []  # Initialize before conditional so it's always defined
            if acct.get("technologies"):
                # Filter to Salesforce-relevant technologies only
                for t in acct["technologies"]:
                    cat = t.get("category", "")
                    name = t.get("name", "")
                    if cat in HIDE_CATEGORIES:
                        continue
                    sf_solution = _get_sf_solution(name, cat)
                    # Skip tools with no Salesforce compete story
                    if sf_solution == "—" or not sf_solution:
                        continue
                    # Determine detection source from the name pattern
                    if "(via SPF)" in name:
                        source = "DNS (SPF)"
                    elif "(via " in name:
                        source = "Subdomain"
                    else:
                        source = "Website"
                    sf_relevant.append({
                        "Category": cat,
                        "Current Tool": name,
                        "SF Solution": sf_solution,
                        "Source": source,
                    })

                # Also add DNS-discovered email/senders
                dns = acct.get("dns", {})
                for provider in dns.get("email_provider", []):
                    sf_relevant.append({
                        "Category": "Email",
                        "Current Tool": provider,
                        "SF Solution": _get_sf_solution(provider, "Email"),
                        "Source": "DNS (MX)",
                    })
                for sender in dns.get("email_senders", []):
                    # Skip if already in the list
                    if not any(r["Current Tool"] == sender for r in sf_relevant):
                        sf_relevant.append({
                            "Category": "Email/Marketing (SPF)",
                            "Current Tool": sender,
                            "SF Solution": _get_sf_solution(sender, "Marketing"),
                            "Source": "DNS (SPF)",
                        })

                if sf_relevant:
                    st.markdown("**Website Detection** _(Wappalyzer + DNS)_")
                    st.markdown(
                        _build_markdown_table(sf_relevant, ["Category", "Current Tool", "SF Solution", "Source"]),
                        unsafe_allow_html=True,
                    )
                else:
                    st.info("No Salesforce-relevant technologies detected on website")

                # Show count of hidden techs
                hidden_count = sum(1 for t in acct["technologies"] if t.get("category", "") in HIDE_CATEGORIES)
                if hidden_count:
                    st.caption(f"_{hidden_count} non-relevant technologies hidden (jQuery, fonts, frameworks, etc.)_")
            else:
                st.info("No website tech stack data available")

            # Job posting tools section
            jp_tools = acct.get("job_posting_tools", [])
            if jp_tools:
                st.markdown("---")
                st.markdown("**Job Posting Discovery** _(tools mentioned in job listings)_")
                jp_rows = []
                for tool in jp_tools:
                    already_detected = any(
                        tool["name"].lower() in r.get("Current Tool", "").lower()
                        for r in sf_relevant
                    )
                    source_label = f"({tool['mention_count']}x)" if tool["mention_count"] > 1 else ""
                    tool_display = f"{'✓ ' if already_detected else ''}{tool['name']} {source_label}".strip()
                    usage_ctx = tool.get("usage_context", "")
                    if usage_ctx:
                        tool_display += f" · _{usage_ctx}_"
                    jp_rows.append({
                        "Category": tool["category"],
                        "Current Tool": tool_display,
                        "SF Solution": tool["sf_solution"],
                    })
                st.markdown(
                    _build_markdown_table(jp_rows, ["Category", "Current Tool", "SF Solution"]),
                    unsafe_allow_html=True,
                )
                # Count new tools not found by website scanning
                new_tools = [t for t in jp_tools if not any(
                    t["name"].lower() in r.get("Current Tool", "").lower()
                    for r in sf_relevant
                )]
                if new_tools:
                    st.caption(f"_🆕 {len(new_tools)} tool(s) only found via job postings — not visible on website_")

        with tab2:
            if acct.get("news_items"):
                tag_colors = {
                    "M&A Activity": "🔴",
                    "Funding/IPO": "🟢",
                    "Major Contract/Deal": "🟡",
                    "Strategic Partnership": "🔵",
                    "Expansion": "🟣",
                    "Digital Transformation": "🟠",
                    "Restructuring": "⚪",
                }

                # Legend
                st.caption(
                    "🔴 M&A  ·  🟢 Funding/IPO  ·  🟡 Contract/Deal  ·  "
                    "🔵 Partnership  ·  🟣 Expansion  ·  🟠 Digital Transformation  ·  ⚪ Restructuring"
                )

                news_rows = []
                for item in acct["news_items"]:
                    events = item.get("events", [])
                    tags = " ".join(tag_colors.get(e, "⚫") for e in events) if events else ""
                    title = item.get("title", "")
                    url = item.get("url", "")
                    date = item.get("date", "")[:10]
                    # Clean summary: strip nav text
                    summary = _clean_summary(item.get("summary", ""))
                    link = f"[{title}]({url})" if url else title
                    news_rows.append({
                        "Date": date,
                        "Signals": tags,
                        "Headline": link,
                        "Summary": summary[:120],
                    })
                st.markdown(
                    _build_markdown_table(news_rows, ["Date", "Signals", "Headline", "Summary"]),
                    unsafe_allow_html=True,
                )
            else:
                st.info("No recent news found")

        with tab3:
            if acct.get("executive_changes"):
                exec_rows = []
                for exc in acct["executive_changes"]:
                    person = exc.get("person_name", "") or "New hire"
                    title = exc.get("title", "")
                    li_url = exc.get("linkedin_url", "")
                    ann_url = exc.get("announcement_url", "")
                    date = exc.get("date", "")[:10]
                    headline = exc.get("headline", "")

                    li_link = f"[Profile]({li_url})" if li_url else "—"
                    ann_link = f"[Article]({ann_url})" if ann_url else "—"

                    exec_rows.append({
                        "Date": date,
                        "Name": person,
                        "Title": title,
                        "LinkedIn": li_link,
                        "Source": ann_link,
                    })
                st.markdown(
                    _build_markdown_table(exec_rows, ["Date", "Name", "Title", "LinkedIn", "Source"]),
                    unsafe_allow_html=True,
                )
            else:
                st.info("No executive changes detected")

        with tab4:
            st.json(acct)
